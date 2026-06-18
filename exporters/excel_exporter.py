"""
excel_exporter.py — Xuất ProjectData ra file Excel (.xlsx) bằng openpyxl.

Bao gồm 3 sheet:
  1. "Task List"   — bảng task chính, indent theo outline_level
  2. "Gantt Chart" — Gantt đơn giản vẽ bằng cách tô màu cell theo timeline
  3. "Resources"   — bảng resource + assignment (chỉ tạo nếu có data)

YÊU CẦU AGENT KHI IMPLEMENT:
- Dùng openpyxl, KHÔNG dùng pandas.to_excel() đơn thuần (cần format
  thủ công: màu nền header, freeze pane, auto-width, highlight milestone).
- Gantt chart: mỗi cột = 1 đơn vị thời gian (ngày hoặc tuần, tự động chọn
  theo độ dài project — < 60 ngày thì theo ngày, ngược lại theo tuần).
  Tô màu cell từ start đến finish của mỗi task trên đúng dòng của nó.
- Task summary (is_summary=True) in đậm, không tô Gantt bar (hoặc tô khác màu).
- Milestone (is_milestone=True): tô màu vàng/cam khác biệt, ghi rõ icon hoặc text "◆".
"""
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.data_model import ProjectData
from exporters.base_exporter import BaseExporter

HEADER_FILL = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_FONT = Font(bold=True)
MILESTONE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GANTT_BAR_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")


class ExcelExporter(BaseExporter):
    def export(self, output_path: str) -> None:
        self.validate_output_path(output_path, ".xlsx")

        wb = Workbook()
        self._build_task_list_sheet(wb.active)
        wb.active.title = "Task List"

        gantt_sheet = wb.create_sheet("Gantt Chart")
        self._build_gantt_sheet(gantt_sheet)

        if self.project_data.resources:
            resource_sheet = wb.create_sheet("Resources")
            self._build_resources_sheet(resource_sheet)

        if self.project_data.assignments:
            assignments_sheet = wb.create_sheet("Assignments")
            self._build_assignments_sheet(assignments_sheet)

        project_info_sheet = wb.create_sheet("Project Info")
        self._build_project_info_sheet(project_info_sheet)

        wb.save(output_path)

    def _build_task_list_sheet(self, sheet) -> None:
        """
        Build sheet "Task List" with task details, indentation, milestone highlights,
        freeze panes, and auto-width formatting.
        """
        headers = [
            "WBS", "Name", "Start", "Finish", "Duration (days)", "% Complete",
            "Critical", "Free Slack", "Total Slack", "Priority", "Cost", "Actual Cost",
            "Constraint", "Predecessors", "Resources", "Notes"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        sheet.freeze_panes = "A2"
        sym = self.project_data.currency_symbol or ""

        for row_idx, task in enumerate(self.project_data.tasks, start=2):
            indent = "  " * task.outline_level
            
            row_values = [
                task.wbs or str(task.outline_level),
                indent + task.name,
                task.start.strftime("%Y-%m-%d") if task.start else "",
                task.finish.strftime("%Y-%m-%d") if task.finish else "",
                task.duration_days,
                task.percent_complete / 100.0,
                "Yes" if task.is_critical else "",
                f"{task.free_slack_days:.2f}d",
                f"{task.total_slack_days:.2f}d",
                str(task.priority),
                task.cost,
                task.actual_cost,
                task.constraint_type.replace("_", " ").title() if task.constraint_type else "",
                ", ".join(str(p) for p in task.predecessors) if task.predecessors else "",
                "; ".join(task.resource_names) if task.resource_names else "",
                task.notes[:32767] if task.notes else "" # Excel max cell length
            ]

            for col_idx, val in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                
                if task.is_summary:
                    cell.font = SUMMARY_FONT
                if task.is_milestone:
                    cell.fill = MILESTONE_FILL

                if col_idx == 5:  # Duration
                    cell.number_format = '0.00'
                elif col_idx == 6:  # % Complete
                    cell.number_format = '0%'
                elif col_idx in (11, 12): # Cost and Actual Cost
                    if val:
                        cell.number_format = f'"{sym}"#,##0'

        # Auto-width formatting
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # Don't size columns on long notes
                if col[0].column == 16 and len(val) > 50:
                    max_len = max(max_len, 50)
                else:
                    val_stripped = val.lstrip()
                    max_len = max(max_len, len(val_stripped))
                    
            col_letter = get_column_letter(col[0].column)
            if col[0].column == 2: # Name
                sheet.column_dimensions[col_letter].width = max(max_len + 12, 30)
            elif col[0].column == 16: # Notes
                sheet.column_dimensions[col_letter].width = 50
            else:
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _build_gantt_sheet(self, sheet) -> None:
        """
        Vẽ Gantt chart bằng cách:
        1. Xác định khoảng thời gian project (start_date -> finish_date)
        2. Chọn đơn vị cột (ngày nếu project < 60 ngày, tuần nếu dài hơn)
        3. Cột A = task name (indent theo outline_level), các cột sau = timeline
        4. Với mỗi task, tô GANTT_BAR_FILL vào các cell tương ứng khoảng start-finish
        5. Milestone: chỉ tô 1 cell duy nhất bằng MILESTONE_FILL tại ngày start
        """
        tasks_with_dates = [t for t in self.project_data.tasks if t.start and t.finish]
        if not tasks_with_dates:
            sheet.cell(row=1, column=1, value="Project không có dữ liệu timeline hợp lệ.")
            return

        proj_start = self.project_data.start_date
        proj_finish = self.project_data.finish_date

        if not proj_start:
            proj_start = min(t.start for t in tasks_with_dates)
        if not proj_finish:
            proj_finish = max(t.finish for t in tasks_with_dates)

        proj_start = proj_start.replace(hour=0, minute=0, second=0, microsecond=0)
        proj_finish = proj_finish.replace(hour=0, minute=0, second=0, microsecond=0)

        total_days = (proj_finish - proj_start).days
        is_daily = total_days < 60

        timeline_cols = []
        if is_daily:
            current = proj_start
            while current <= proj_finish:
                timeline_cols.append(current)
                current += timedelta(days=1)
        else:
            start_monday = proj_start - timedelta(days=proj_start.weekday())
            current = start_monday
            while current <= proj_finish:
                timeline_cols.append(current)
                current += timedelta(weeks=1)

        cell_a1 = sheet.cell(row=1, column=1, value="Task Name")
        cell_a1.fill = HEADER_FILL
        cell_a1.font = HEADER_FONT

        for col_idx, date_val in enumerate(timeline_cols, start=2):
            cell = sheet.cell(row=1, column=col_idx)
            if is_daily:
                cell.value = date_val.strftime("%m-%d")
            else:
                cell.value = f"W{date_val.isocalendar()[1]}/{date_val.strftime('%m-%d')}"
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        SUMMARY_BAR_FILL = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        for row_idx, task in enumerate(self.project_data.tasks, start=2):
            indent = "  " * task.outline_level
            cell_name = sheet.cell(row=row_idx, column=1, value=indent + task.name)
            if task.is_summary:
                cell_name.font = SUMMARY_FONT

            if not task.start or not task.finish:
                continue

            t_start = task.start.replace(hour=0, minute=0, second=0, microsecond=0)
            t_finish = task.finish.replace(hour=0, minute=0, second=0, microsecond=0)

            for col_idx, col_date in enumerate(timeline_cols, start=2):
                cell = sheet.cell(row=row_idx, column=col_idx)

                if is_daily:
                    if t_start <= col_date <= t_finish:
                        if task.is_milestone:
                            cell.value = "◆"
                            cell.fill = MILESTONE_FILL
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")
                        elif task.is_summary:
                            cell.fill = SUMMARY_BAR_FILL
                        else:
                            cell.fill = GANTT_BAR_FILL
                else:
                    week_end = col_date + timedelta(days=6)
                    if max(t_start, col_date) <= min(t_finish, week_end):
                        if task.is_milestone:
                            cell.value = "◆"
                            cell.fill = MILESTONE_FILL
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")
                        elif task.is_summary:
                            cell.fill = SUMMARY_BAR_FILL
                        else:
                            cell.fill = GANTT_BAR_FILL

        max_len = 0
        for row in range(1, len(self.project_data.tasks) + 2):
            val = str(sheet.cell(row=row, column=1).value or '')
            val_stripped = val.lstrip()
            max_len = max(max_len, len(val_stripped))
        sheet.column_dimensions["A"].width = max(max_len + 12, 20)

        for col_idx in range(2, len(timeline_cols) + 2):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 10

    def _build_resources_sheet(self, sheet) -> None:
        headers = ["ID", "Resource Name", "Type", "Initials", "Max Units",
                   "Calendar", "Standard Rate", "Email", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        sheet.freeze_panes = "A2"

        for row_idx, resource in enumerate(self.project_data.resources, start=2):
            row_values = [
                str(resource.id),
                resource.name,
                resource.type,
                resource.initials,
                resource.max_units / 100.0 if resource.max_units is not None else None,
                resource.calendar_name,
                resource.std_rate,
                resource.email,
                resource.notes[:32767] if resource.notes else ""
            ]

            for col_idx, val in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 5: # Max Units
                    cell.number_format = '0%'

        self._auto_fit_columns(sheet)

    def _build_assignments_sheet(self, sheet) -> None:
        headers = ["Task ID", "Task Name", "Resource ID", "Resource Name",
                   "Units", "Work (h)", "Actual Work (h)", "Cost", "Actual Cost"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        sheet.freeze_panes = "A2"

        task_map = {t.id: t.name for t in self.project_data.tasks}
        res_map = {r.id: r.name for r in self.project_data.resources}
        sym = self.project_data.currency_symbol or ""

        for row_idx, a in enumerate(self.project_data.assignments, start=2):
            task_name = task_map.get(a.task_id, "")
            res_name = res_map.get(a.resource_id, "")
            
            row_values = [
                str(a.task_id), task_name,
                str(a.resource_id), res_name,
                a.units / 100.0 if a.units is not None else None,
                a.work_hours,
                a.actual_work_hours,
                a.cost,
                a.actual_cost
            ]
            
            for col_idx, val in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 5:
                    cell.number_format = '0%'
                elif col_idx in (6, 7):
                    if val is not None:
                        cell.number_format = '0.0"h"'
                elif col_idx in (8, 9):
                    if val is not None:
                        cell.number_format = f'"{sym}"#,##0'

        self._auto_fit_columns(sheet)

    def _build_project_info_sheet(self, sheet) -> None:
        data = self.project_data
        
        # Standard Properties
        fields = [
            ("Tên project",          data.name),
            ("Tác giả (Author)",      data.author or "—"),
            ("Ngày bắt đầu",          data.start_date.strftime("%Y-%m-%d %H:%M") if data.start_date else "—"),
            ("Ngày kết thúc",          data.finish_date.strftime("%Y-%m-%d %H:%M") if data.finish_date else "—"),
            ("Lần lưu cuối",           data.last_saved.strftime("%Y-%m-%d %H:%M") if data.last_saved else "—"),
            ("Lịch sửa (Revision)",   str(data.revision)),
            ("Tiền tệ",               data.currency_symbol or "—"),
            ("Giờ làm việc/ngày",    f"{data.minutes_per_day // 60}h"),
            ("Calendar mặc định",     data.calendar_name or "—"),
            ("Tổng tasks",             str(data.task_count)),
            ("Milestones",             str(data.milestone_count)),
            ("Resources",              str(len(data.resources))),
            ("Assignments",            str(len(data.assignments))),
        ]

        # Header Std
        sheet.cell(row=1, column=1, value="Project Properties").font = SUMMARY_FONT
        sheet.cell(row=1, column=1).fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        
        row_idx = 2
        for lbl, val in fields:
            sheet.cell(row=row_idx, column=1, value=lbl).font = SUMMARY_FONT
            sheet.cell(row=row_idx, column=2, value=val)
            row_idx += 1
            
        row_idx += 1
        
        # Custom Properties
        sheet.cell(row=row_idx, column=1, value="Custom Properties").font = SUMMARY_FONT
        sheet.cell(row=row_idx, column=1).fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        row_idx += 1
        
        sheet.cell(row=row_idx, column=1, value="Property").font = SUMMARY_FONT
        sheet.cell(row=row_idx, column=2, value="Value").font = SUMMARY_FONT
        row_idx += 1
        
        for key, val in sorted(data.custom_properties.items()):
            sheet.cell(row=row_idx, column=1, value=key)
            sheet.cell(row=row_idx, column=2, value=val)
            row_idx += 1

        self._auto_fit_columns(sheet)

    def _auto_fit_columns(self, sheet) -> None:
        """Helper để tự động dãn cột cho một sheet bất kỳ"""
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # Giới hạn với các dòng text quá dài (notes)
                if len(val) > 50:
                    max_len = max(max_len, 50)
                else:
                    max_len = max(max_len, len(val))
            
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

